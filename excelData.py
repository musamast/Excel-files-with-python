#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
wb = openpyxl.load_workbook('CS remaining products.xlsx')


# In[2]:


sheet = wb['Sheet1']


# In[3]:


import string
newWord=[]
for i in range(2,sheet.max_row+1):
    txt=sheet.cell(i,4).value
    if (txt is not None):
        wordList=txt.split()
        for word in wordList:
            if (word.endswith("/cs")) or (word.endswith("/CS")):
                quantity=wordList[wordList.index(word)-1]
                if "," in quantity:
                    newWord=quantity.split(',')
                    sheet.cell(i,5).value=newWord[1]
                else:
                    sheet.cell(i,5).value=quantity
                neww=''
                for di in word:
                    if di in string.digits:
                        neww+=di
                        sheet.cell(i,5).value=neww
            


# In[4]:


wb.save("C:\\Users\\user\\Desktop\\demo23.xlsx")


# In[ ]:


import string
newWord=[]
for i in range(2,sheet.max_row+1):
    txt=sheet.cell(i,3).value
    if (txt is not None):
        wordList=txt.split()
        for word in wordList:
            if (word.endswith("air")) or (word.endswith("pr")):
                quantity=wordList[wordList.index(word)-1]
                if "," in quantity:
                    newWord=quantity.split(',')
                    sheet.cell(i,8).value=newWord[1]
                else:
                    sheet.cell(i,8).value=quantity
                neww=''
                for di in word:
                    if di in string.digits:
                        neww+=di
                        sheet.cell(i,8).value=neww
            


# In[ ]:




